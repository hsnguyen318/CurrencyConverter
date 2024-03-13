from openpyxl import load_workbook
import datetime
import xlrd
from flask import Flask, request

currency_list = {'AUD': 1, 'EUR': 2, 'NZD': 3, 'GBP': 4, 'BRL': 5, 'CAD': 6, 'CNY': 7, 'DKK': 8, 'HKD': 9, 'INR': 10,
                 'JPY': 11, 'MYR': 12, 'MXN': 13, 'NOK': 14, 'ZAR': 15, 'SGD': 16, 'KRW': 17, 'LKR': 18, 'SEK': 19,
                 'CHF': 20, 'TWD': 21, 'THB': 22, 'VEB': 23}


def get_data(from_curr, to_curr, from_date, to_date):
    result = []
    # read by default 1st sheet of an excel file
    data_file = 'FXdatabase.xlsx'
    wb = load_workbook(data_file, read_only=True)
    ws = wb['Sheet1']
    all_rows = list(ws.rows)

    # calculate range
    from_date_value = datetime.date(int(from_date[0:4]), int(from_date[5:7]), int(from_date[8:10]))
    to_date_value = datetime.date(int(to_date[0:4]), int(to_date[5:7]), int(to_date[8:10]))
    delta = to_date_value - from_date_value

    start = 0
    workbook = xlrd.open_workbook('FXdatabase.xls')
    sheet = workbook['Sheet1']

    for row in range(sheet.nrows):
        # print("row::::: ", row)
        if sheet.cell(row, 0).value == from_date:
            start = row
    for num in range(start, start + delta.days + 1, 1):
        if from_curr != 'USD' and to_curr != 'USD':
            result.append((sheet.cell(start, 0).value,
                           sheet.cell(start, currency_list[to_curr]).value/sheet.cell(start, currency_list[from_curr]).value))

        elif from_curr == 'USD' and to_curr != 'USD':
            result.append((sheet.cell(start, 0).value,
                       sheet.cell(start, currency_list[to_curr]).value))

        elif from_curr != 'USD' and to_curr == 'USD':
            result.append((sheet.cell(start, 0).value,
                           1 / sheet.cell(start, currency_list[to_curr]).value))
        else:
            result.append((sheet.cell(start, 0).value,1.0))
        start += 1
    return result


# function to return data to graph making microservice
app = Flask(__name__)


@app.route('/api', methods=['GET'])


def my_api():
    from_curr = request.args.get('from_curr')
    to_curr = request.args.get('to_curr')
    from_date = request.args.get('from_date')
    to_date = request.args.get('to_date')

    # do something with the parameters...

    return get_data(from_curr, to_curr, from_date, to_date)


if __name__ == '__main__':
    app.run(debug=True)
